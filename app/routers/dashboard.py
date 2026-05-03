from __future__ import annotations
import io
from datetime import UTC, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..data_engine import load_market_status
from ..database import get_db
from ..ml_model import load_metrics
from ..models import AnalysisResult, Stock, User, WatchlistItem
from ..scoring import RANKING_EMOJI

router    = APIRouter()
templates = Jinja2Templates(directory="templates")


def _compute_target_price(close: float | None, atr: float | None, wranking: str, hg_score: int | None) -> float | None:
    if not close or not atr:
        return None
    if wranking == "Strong Buy":
        n = 3.0 if hg_score is not None else 2.5
    elif wranking == "Buy":
        n = 2.0 if hg_score is not None else 1.5
    else:
        return None
    return round(close + n * atr, 2)


def _compute_weighted(score_final: float, fundamental_score, wt: int, wf: int) -> tuple[int, str]:
    tech = score_final or 0
    if wf == 0 or fundamental_score is None:
        wscore = tech
    elif wt == 0:
        wscore = fundamental_score
    else:
        wscore = round(tech * wt / 100 + fundamental_score * wf / 100)
    wscore = int(wscore)
    if wscore >= 75:
        return wscore, "Strong Buy"
    if wscore >= 58:
        return wscore, "Buy"
    if wscore >= 42:
        return wscore, "Neutral"
    return wscore, "Avoid"


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    market: str  = Query("CAC40"),
    ranking: str = Query(""),
    sector: str  = Query(""),
    weight: str  = Query("65-35"),
    hyper_growth: str = Query(""),   # "1" pour ne montrer que les 🦄 éligibles
    watchlist: str = Query(""),      # "1" pour ne montrer que les tickers en watchlist
    db: Session  = Depends(get_db),
    user: User   = Depends(get_current_user),
):
    try:
        wt, wf = [int(x) for x in weight.split("-")]
        if wt + wf != 100:
            wt, wf = 65, 35
    except Exception:
        wt, wf = 65, 35
    weight = f"{wt}-{wf}"

    last_date = (
        db.query(AnalysisResult.date)
        .order_by(AnalysisResult.date.desc())
        .limit(1)
        .scalar()
    )
    if last_date is None:
        results = []
        user_watchlist: set[str] = set()
    else:
        query = (
            db.query(AnalysisResult, Stock)
            .join(Stock, AnalysisResult.stock_id == Stock.id)
            .filter(AnalysisResult.date == last_date)
            .filter(Stock.is_active.is_(True))
        )
        if market:
            query = query.filter(Stock.market == market)
        if ranking:
            query = query.filter(AnalysisResult.ranking == ranking)
        if sector:
            query = query.filter(Stock.sector == sector)
        if hyper_growth == "1":
            query = query.filter(AnalysisResult.hyper_growth_score.isnot(None))

        # Watchlist filter
        user_watchlist: set[str] = {
            w.ticker for w in
            db.query(WatchlistItem.ticker).filter(WatchlistItem.user_id == user.id).all()
        }
        if watchlist == "1":
            query = query.filter(Stock.ticker.in_(user_watchlist))

        rows = query.limit(2500).all()

        results = []
        for ar, stock in rows:
            wscore, wranking = _compute_weighted(ar.score_final, ar.fundamental_score, wt, wf)
            results.append({
                "ticker":             stock.ticker,
                "name":               stock.name or "",
                "market":             stock.market,
                "sector":             stock.sector or "",
                "close":              round(ar.close or 0, 2),
                "score_final":        ar.score_final or 0,
                "score_composite":    ar.score_composite,
                "fundamental_score":  ar.fundamental_score,
                "ranking":            ar.ranking or "Neutral",
                "weighted_score":     wscore,
                "weighted_ranking":   wranking,
                "weighted_emoji":     RANKING_EMOJI.get(wranking, "⚪"),
                "rsi":                round(ar.rsi or 0, 1),
                "macd_hist":          round(ar.macd_hist or 0, 4),
                "macd_bull":          (ar.macd_hist or 0) > 0,
                "stop_loss":          round(ar.stop_loss_price or 0, 2),
                "volatility":         round(ar.volatility or 0, 1),
                "ml_prob":            round(ar.ml_probability * 100, 1) if ar.ml_probability else None,
                "bollinger_b":        round(ar.bollinger_b or 0, 2),
                "atr_pct":            round((ar.atr / ar.close * 100) if ar.atr and ar.close else 0, 2),
                "adx":                round(ar.adx or 0, 1) if ar.adx is not None else None,
                "sma200_slope":       round(ar.sma200_slope or 0, 2) if ar.sma200_slope is not None else None,
                "atr_pct_rank":       round(ar.atr_pct_rank or 0, 0) if ar.atr_pct_rank is not None else None,
                "bb_zscore":          round(ar.bb_zscore or 0, 2) if ar.bb_zscore is not None else None,
                "pe_ratio":           ar.pe_ratio,
                "pb_ratio":           ar.pb_ratio,
                "roe":                ar.roe,
                "debt_equity":        round(ar.debt_equity / 100, 2) if ar.debt_equity else None,
                "rev_growth":         ar.rev_growth,
                "peg_ratio":          round(ar.peg_ratio, 2)  if ar.peg_ratio  else None,
                "ev_ebit":            round(ar.ev_ebit, 1)    if ar.ev_ebit    else None,
                "ev_ebitda":          round(ar.ev_ebitda, 1)  if ar.ev_ebitda  else None,
                "fcf":                ar.fcf,
                "hyper_growth_score": ar.hyper_growth_score,
                "news_sentiment":     ar.news_sentiment,
                "target_price":       _compute_target_price(ar.close, ar.atr, wranking, ar.hyper_growth_score),
                "watching":           stock.ticker in user_watchlist,
            })

        if hyper_growth == "1":
            # Tri spécifique : par hyper_growth_score desc
            results.sort(key=lambda x: x["hyper_growth_score"] or 0, reverse=True)
        else:
            results.sort(key=lambda x: x["weighted_score"], reverse=True)

    # Secteurs disponibles pour le marché sélectionné (ou tous si pas de marché)
    sector_q = db.query(Stock.sector).filter(Stock.sector.isnot(None), Stock.sector != "")
    if market:
        sector_q = sector_q.filter(Stock.market == market)
    sectors = sorted({r[0] for r in sector_q.distinct().all()})

    ml_metrics    = load_metrics()
    market_status = load_market_status()
    markets       = ["CAC40", "SBF120", "EURONEXT_GROWTH", "SP500", "NASDAQ", "COMMODITIES", "CRYPTO"]
    rankings      = ["Strong Buy", "Buy", "Neutral", "Avoid"]
    last_update   = last_date.strftime("%d/%m/%Y") if last_date else "Aucune donnée"
    quote_status  = market_status.get(market, {}).get("display")
    market_state  = market_status.get(market, {}).get("market_state", "")
    is_open       = market_state == "REGULAR"

    return templates.TemplateResponse(request, "dashboard.html", {
        "user":         user,
        "results":      results,
        "markets":      markets,
        "rankings":     rankings,
        "sectors":      sectors,
        "sel_market":   market,
        "sel_ranking":  ranking,
        "sel_sector":   sector,
        "sel_weight":   weight,
        "hyper_growth_filter": hyper_growth,
        "watchlist_filter": watchlist,
        "last_update":  last_update,
        "ml_metrics":   ml_metrics,
        "quote_status": quote_status,
        "is_open":      is_open,
    })


@router.get("/dashboard/export")
def export_excel(
    market:  str   = Query("CAC40"),
    tickers: str   = Query(""),        # comma-separated, empty = all
    db: Session    = Depends(get_db),
    user: User     = Depends(get_current_user),
):
    last_date = (
        db.query(AnalysisResult.date)
        .order_by(AnalysisResult.date.desc())
        .limit(1)
        .scalar()
    )
    if not last_date:
        return HTMLResponse("Aucune donnée", status_code=404)

    ticker_filter = [t.strip().upper() for t in tickers.split(",") if t.strip()]
    query = (
        db.query(AnalysisResult, Stock)
        .join(Stock, AnalysisResult.stock_id == Stock.id)
        .filter(AnalysisResult.date == last_date)
        .filter(Stock.is_active.is_(True))
    )
    if ticker_filter:
        query = query.filter(Stock.ticker.in_(ticker_filter))
    else:
        query = query.filter(Stock.market == market)

    rows = query.order_by(AnalysisResult.score_final.desc()).all()

    # ── Build xlsx ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = market

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="E2E8F0")
    green_font  = Font(color="4ADE80", bold=True)
    red_font    = Font(color="F87171", bold=True)

    headers = [
        "Ticker", "Nom", "Marché", "Secteur", "Prix", "Score Final", "Score Composite",
        "Score Fond.", "Signal", "RSI", "MACD ▲▼", "Volatilité %",
        "Stop-Loss", "ML Prob %", "ATR %", "Bollinger %B",
        "P/E", "P/B", "ROE %", "D/E", "Croiss. Rev %",
        "PEG", "EV/EBIT", "EV/EBITDA", "FCF",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ar, stock in rows:
        de  = round(ar.debt_equity / 100, 2) if ar.debt_equity else None
        row = [
            stock.ticker,
            stock.name or "",
            stock.market,
            stock.sector or "",
            round(ar.close or 0, 2),
            ar.score_final or 0,
            ar.score_composite,
            ar.fundamental_score,
            ar.ranking or "Neutral",
            round(ar.rsi or 0, 1),
            "▲" if (ar.macd_hist or 0) > 0 else "▼",
            round(ar.volatility or 0, 1),
            round(ar.stop_loss_price or 0, 2),
            round((ar.ml_probability or 0) * 100, 1),
            round((ar.atr / ar.close * 100) if ar.atr and ar.close else 0, 2),
            round(ar.bollinger_b or 0, 2),
            ar.pe_ratio,
            ar.pb_ratio,
            round(ar.roe * 100, 1) if ar.roe else None,
            de,
            round(ar.rev_growth * 100, 1) if ar.rev_growth else None,
            round(ar.peg_ratio, 2)  if ar.peg_ratio  else None,
            round(ar.ev_ebit, 1)    if ar.ev_ebit    else None,
            round(ar.ev_ebitda, 1)  if ar.ev_ebitda  else None,
            ar.fcf,
        ]
        ws.append(row)
        r = ws.max_row
        signal_cell = ws.cell(r, 9)
        if ar.ranking in ("Strong Buy", "Buy"):
            signal_cell.font = green_font
        elif ar.ranking == "Avoid":
            signal_cell.font = red_font
        score_cell = ws.cell(r, 6)
        if (ar.score_final or 0) >= 75:
            score_cell.font = green_font
        elif (ar.score_final or 0) < 42:
            score_cell.font = red_font

    col_widths = [10, 28, 12, 20, 10, 12, 14, 12, 12, 8, 10, 12,
                  11, 10, 8, 12, 8, 8, 8, 8, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = last_date.strftime("%Y-%m-%d")
    filename = f"StockAnalyzer_{market}_{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/hyper-growth-explain/{ticker}")
def hyper_growth_explain(
    ticker: str,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from fastapi.responses import JSONResponse
    from ..models import AnalysisResult, Stock

    stock = db.query(Stock).filter(Stock.ticker == ticker).first()
    if not stock:
        return JSONResponse({"error": "Ticker introuvable"}, status_code=404)

    ar = (
        db.query(AnalysisResult)
        .filter(AnalysisResult.stock_id == stock.id)
        .order_by(AnalysisResult.date.desc())
        .first()
    )
    if not ar:
        return JSONResponse({"error": "Pas de données"}, status_code=404)

    rev_growth      = ar.rev_growth        # % (ex: 25.0)
    debt_equity     = ar.debt_equity       # ×100 (ex: 150 = 1.5×)
    fundamental_score = ar.fundamental_score
    score_final     = ar.score_final or 0
    fcf             = ar.fcf
    pb_ratio        = ar.pb_ratio
    obv_slope       = ar.volume_ratio      # proxy : volume_ratio reflète l'OBV trend

    # ── Éligibilité ──────────────────────────────────────────────────────────
    ok_growth = rev_growth is not None and rev_growth >= 15
    ok_obv    = (ar.volume_ratio or 0) >= 1.0   # volume ratio comme proxy OBV
    ok_score  = score_final >= 55
    ok_fonda  = fundamental_score is not None
    ok_de     = debt_equity is None or debt_equity <= 300

    de_display = f"{debt_equity/100:.1f}×" if debt_equity is not None else "N/A"

    eligibility = [
        {"label": "Croissance revenus ≥ 15%",
         "value": f"{rev_growth:.1f}%" if rev_growth is not None else "N/A",
         "pass": ok_growth},
        {"label": "Volume ratio ≥ 1.0 (accumulation)",
         "value": f"{ar.volume_ratio:.2f}×" if ar.volume_ratio else "N/A",
         "pass": ok_obv},
        {"label": "Score technique ≥ 55",
         "value": str(score_final),
         "pass": ok_score},
        {"label": "Fondamentaux disponibles",
         "value": "Oui" if ok_fonda else "Non",
         "pass": ok_fonda},
        {"label": "D/E ≤ 3×",
         "value": de_display,
         "pass": ok_de},
    ]
    eligible = all(e["pass"] for e in eligibility)

    # ── Décomposition du score ────────────────────────────────────────────────
    breakdown = []
    if eligible:
        # 1. Croissance (40 pts)
        if   rev_growth >= 50: g_pts, g_detail = 40, "≥ 50%"
        elif rev_growth >= 30: g_pts, g_detail = 35, "≥ 30%"
        elif rev_growth >= 25: g_pts, g_detail = 30, "≥ 25%"
        elif rev_growth >= 20: g_pts, g_detail = 25, "≥ 20%"
        else:                  g_pts, g_detail = 18, "15–20%"
        breakdown.append({"label": "Croissance revenus", "pts": g_pts, "max": 40, "detail": g_detail})

        # 2. Momentum technique (25 pts)
        adx   = ar.adx or 0
        slope = ar.sma200_slope or 0
        t_pts = 0
        t_parts = []
        if   adx >= 30: t_pts += 10; t_parts.append(f"ADX {adx:.0f} (fort)")
        elif adx >= 25: t_pts += 7;  t_parts.append(f"ADX {adx:.0f} (moyen)")
        elif adx >= 20: t_pts += 4;  t_parts.append(f"ADX {adx:.0f} (faible)")
        else:           t_parts.append(f"ADX {adx:.0f} (range)")
        if   slope > 0.5: t_pts += 10; t_parts.append("SMA200 ↑↑")
        elif slope > 0:   t_pts += 6;  t_parts.append("SMA200 ↑")
        else:             t_parts.append("SMA200 ↓")
        t_pts += 5  # OBV slope > 0 déjà vérifié en éligibilité
        t_parts.append("OBV ✓")
        t_pts = min(25, t_pts)
        breakdown.append({"label": "Momentum technique", "pts": t_pts, "max": 25,
                          "detail": " · ".join(t_parts)})

        # 3. Accumulation volume (20 pts)
        vr = ar.volume_ratio or 1
        v_pts = 0
        v_parts = []
        if   vr >= 1.5: v_pts += 12; v_parts.append(f"Vol ratio {vr:.2f}× (fort)")
        elif vr >= 1.3: v_pts += 8;  v_parts.append(f"Vol ratio {vr:.2f}× (bon)")
        elif vr >= 1.0: v_pts += 4;  v_parts.append(f"Vol ratio {vr:.2f}× (neutre)")
        v_pts += 5; v_parts.append("OBV positif")  # obv_slope > 0
        v_pts = min(20, v_pts)
        breakdown.append({"label": "Accumulation volume", "pts": v_pts, "max": 20,
                          "detail": " · ".join(v_parts)})

        # 4. Valorisation / FCF (15 pts)
        val_pts = 0
        val_parts = []
        if fcf is not None and fcf > 0:
            val_pts += 8; val_parts.append(f"FCF positif")
        elif fcf is None:
            val_pts += 4; val_parts.append("FCF N/A")
        else:
            val_parts.append("FCF négatif")
        if pb_ratio is not None:
            if   pb_ratio < 5 and rev_growth > 25: val_pts += 7; val_parts.append(f"P/B {pb_ratio:.1f} (raisonnable)")
            elif pb_ratio < 10:                    val_pts += 4; val_parts.append(f"P/B {pb_ratio:.1f}")
            else:                                  val_parts.append(f"P/B {pb_ratio:.1f} (élevé)")
        else:
            val_pts += 4; val_parts.append("P/B N/A")
        val_pts = min(15, val_pts)
        breakdown.append({"label": "Valorisation / FCF", "pts": val_pts, "max": 15,
                          "detail": " · ".join(val_parts)})

    return JSONResponse({
        "ticker":            ticker,
        "name":              stock.name or ticker,
        "hyper_growth_score": ar.hyper_growth_score,
        "eligible":          eligible,
        "eligibility":       eligibility,
        "breakdown":         breakdown,
    })


@router.get("/api/score-explain/{ticker}")
def score_explain(
    ticker: str,
    weight: str = Query("65-35"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from fastapi.responses import JSONResponse
    from ..fundamentals import _score_pe, _score_pb, _score_roe, _score_debt, _score_growth

    stock = db.query(Stock).filter(Stock.ticker == ticker.upper()).first()
    if not stock:
        return JSONResponse({"error": "Ticker introuvable"}, status_code=404)
    ar = (
        db.query(AnalysisResult)
        .filter(AnalysisResult.stock_id == stock.id)
        .order_by(AnalysisResult.date.desc())
        .first()
    )
    if not ar:
        return JSONResponse({"error": "Pas de données"}, status_code=404)

    try:
        wt, wf = [int(x) for x in weight.split("-")]
        if wt + wf != 100:
            wt, wf = 65, 35
    except Exception:
        wt, wf = 65, 35

    close      = ar.close or 0
    ema50      = ar.ema50 or 0
    sma200     = ar.sma200 or 0
    rsi        = ar.rsi or 50
    macd_hist  = ar.macd_hist or 0
    vol_ratio  = ar.volume_ratio or 1
    adx        = ar.adx or 25
    slope      = ar.sma200_slope or 0
    ml_prob    = ar.ml_probability
    ml_boost   = ar.ml_boost or 0
    score_base = ar.score_base or 0
    score_final = ar.score_final or 0

    # ── Tendance (25 pts) ─────────────────────────────────────────────────────
    t_pts, t_items = 0, []
    if ema50 > sma200:
        t_pts += 10; t_items.append({"label": "Golden cross (EMA50 > SMA200)", "pts": "+10", "ok": True})
    else:
        t_items.append({"label": "Pas de golden cross (EMA50 ≤ SMA200)", "pts": "+0", "ok": False})
    if close > ema50:
        t_pts += 8; t_items.append({"label": f"Prix {close:.2f} > EMA50 {ema50:.2f}", "pts": "+8", "ok": True})
    else:
        t_items.append({"label": f"Prix {close:.2f} < EMA50 {ema50:.2f}", "pts": "+0", "ok": False})
    if close > sma200:
        t_pts += 7; t_items.append({"label": f"Prix > SMA200 {sma200:.2f}", "pts": "+7", "ok": True})
    else:
        t_items.append({"label": f"Prix < SMA200 {sma200:.2f}", "pts": "+0", "ok": False})

    # ── Momentum (25 pts) ─────────────────────────────────────────────────────
    m_pts, m_items = 0, []
    if 50 <= rsi <= 70:
        m_pts += 12; m_items.append({"label": f"RSI {rsi:.0f} — zone haussière (50-70)", "pts": "+12", "ok": True})
    elif rsi < 30:
        if macd_hist > 0:
            m_pts += 8; m_items.append({"label": f"RSI {rsi:.0f} — survente + rebond MACD", "pts": "+8", "ok": True})
        else:
            m_items.append({"label": f"RSI {rsi:.0f} — survente sans confirmation MACD (couteau)", "pts": "+0", "ok": False})
    elif rsi < 50:
        m_pts += 5; m_items.append({"label": f"RSI {rsi:.0f} — zone neutre/basse", "pts": "+5", "ok": None})
    else:
        m_pts -= 5; m_items.append({"label": f"RSI {rsi:.0f} — surachat, risque correction", "pts": "−5", "ok": False})
    if macd_hist > 0:
        m_pts += 13; m_items.append({"label": f"MACD histogramme positif ({macd_hist:.4f})", "pts": "+13", "ok": True})
    else:
        m_items.append({"label": f"MACD histogramme négatif ({macd_hist:.4f})", "pts": "+0", "ok": False})

    # ── Volume (10 pts via ratio) ─────────────────────────────────────────────
    v_pts, v_items = 0, []
    if vol_ratio >= 1.5:
        v_pts += 10; v_items.append({"label": f"Volume x{vol_ratio:.2f} — très fort", "pts": "+10", "ok": True})
    elif vol_ratio >= 1.3:
        v_pts += 5; v_items.append({"label": f"Volume x{vol_ratio:.2f} — fort", "pts": "+5", "ok": True})
    else:
        v_items.append({"label": f"Volume x{vol_ratio:.2f} — normal", "pts": "+0", "ok": None})

    # ── Régimes (-10 à +5) ────────────────────────────────────────────────────
    r_pts, r_items = 0, []
    if adx < 20:
        r_pts -= 5; r_items.append({"label": f"ADX {adx:.0f} — marché en range (−5)", "pts": "−5", "ok": False})
    elif adx > 30:
        r_pts += 3; r_items.append({"label": f"ADX {adx:.0f} — tendance forte (+3)", "pts": "+3", "ok": True})
    else:
        r_items.append({"label": f"ADX {adx:.0f} — tendance modérée", "pts": "+0", "ok": None})
    if slope < 0:
        r_pts -= 5; r_items.append({"label": f"SMA200 baissière ({slope:.2f}%) — biais négatif", "pts": "−5", "ok": False})
    else:
        r_items.append({"label": f"SMA200 haussière ({slope:.2f}%)", "pts": "+0", "ok": True})

    # OBV + Ichimoku résiduel (non stockés individuellement)
    residual = max(0, score_base - t_pts - m_pts - v_pts - r_pts)
    obv_est  = min(10, residual)
    ich_est  = min(15, max(0, residual - 10))

    # ── ML Boost ─────────────────────────────────────────────────────────────
    ml_items = []
    if ml_prob is None:
        ml_items.append({"label": "Modèle ML non disponible", "pts": "+0", "ok": None})
    elif ml_prob >= 0.70:
        ml_items.append({"label": f"Probabilité {ml_prob*100:.0f}% — très forte conviction", "pts": "+15", "ok": True})
    elif ml_prob >= 0.60:
        ml_items.append({"label": f"Probabilité {ml_prob*100:.0f}% — bonne conviction", "pts": "+8", "ok": True})
    elif ml_prob >= 0.50:
        ml_items.append({"label": f"Probabilité {ml_prob*100:.0f}% — légère conviction", "pts": "+3", "ok": None})
    elif ml_prob >= 0.40:
        ml_items.append({"label": f"Probabilité {ml_prob*100:.0f}% — neutre", "pts": "+0", "ok": None})
    else:
        ml_items.append({"label": f"Probabilité {ml_prob*100:.0f}% — signal négatif", "pts": "−15", "ok": False})

    tech_sections = [
        {"title": "📈 Tendance",        "pts": t_pts,    "max": 25, "items": t_items},
        {"title": "⚡ Momentum",         "pts": m_pts,    "max": 25, "items": m_items},
        {"title": "📊 Volume (ratio)",   "pts": v_pts,    "max": 10, "items": v_items},
        {"title": "🔁 OBV (estimé)",     "pts": obv_est,  "max": 10, "items": [{"label": "OBV slope non stocké — estimé depuis score base", "pts": f"+{obv_est}", "ok": obv_est > 0 or None}]},
        {"title": "☁️ Ichimoku (estimé)","pts": ich_est,  "max": 15, "items": [{"label": "Ichimoku non stocké — estimé depuis score base", "pts": f"+{ich_est}", "ok": ich_est > 5 or None}]},
        {"title": "🌍 Régimes marché",   "pts": r_pts,    "max": 5,  "items": r_items},
        {"title": "🤖 Boost ML",         "pts": ml_boost, "max": 15, "items": ml_items},
    ]

    # ── Fondamentaux ──────────────────────────────────────────────────────────
    funda_sections = []
    if ar.fundamental_score is not None:
        pe     = ar.pe_ratio
        pb     = ar.pb_ratio
        roe_f  = ar.roe / 100    if ar.roe       is not None else None
        de     = ar.debt_equity
        grow_f = ar.rev_growth / 100 if ar.rev_growth is not None else None
        funda_sections = [
            {"title": "P/E Ratio",         "pts": _score_pe(pe),      "max": 30, "detail": f"{pe:.1f}"              if pe       else "N/A"},
            {"title": "P/B Ratio",         "pts": _score_pb(pb),      "max": 20, "detail": f"{pb:.2f}"              if pb       else "N/A"},
            {"title": "ROE",               "pts": _score_roe(roe_f),  "max": 25, "detail": f"{ar.roe:.1f}%"         if ar.roe   is not None else "N/A"},
            {"title": "Dette / Capitaux",  "pts": _score_debt(de),    "max": 15, "detail": f"{de/100:.1f}×"         if de       is not None else "N/A"},
            {"title": "Croissance revenus","pts": _score_growth(grow_f),"max":10, "detail": f"{ar.rev_growth:.1f}%" if ar.rev_growth is not None else "N/A"},
        ]

    wscore, wranking = _compute_weighted(score_final, ar.fundamental_score, wt, wf)
    target = _compute_target_price(ar.close, ar.atr, wranking, ar.hyper_growth_score)

    return JSONResponse({
        "ticker":           ticker.upper(),
        "name":             stock.name or ticker,
        "weight":           f"{wt}/{wf}",
        "wt": wt, "wf": wf,
        "score_final":      score_final,
        "score_base":       score_base,
        "ml_boost":         ml_boost,
        "fundamental_score": ar.fundamental_score,
        "weighted_score":   wscore,
        "weighted_ranking": wranking,
        "target_price":     target,
        "close":            round(ar.close or 0, 2),
        "tech_sections":    tech_sections,
        "funda_sections":   funda_sections,
    })
